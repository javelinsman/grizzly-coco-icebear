import openpyxl
import json
import os
import re

nicknames = [
        'iliiiillllliii', #P0, 83
        '굴탱', #P1, 35
        '생애첫행복', #P2, 22
        '예비단비맘', #P3, 47
        '숌찡', #P4, 19
        '줌마양이', #P5, 141
        "현명콩", #P6, 90
        "뱀소띠맘", #P7, 16
        "달짜기", #P8, 60
        "심쿵9", #P9, 12
        "여름은랄라", #P10, 56
        "별베맘", #P11, 29
        "jjjjjjiiieun", #P12, 11
        "가각", #P13, 76
        "hyojung2", #P14, 28
        "야옹이가멍멍해", #P15, 161
        "오늘두시", #P16, 31
        "빵떠기", #P17, 39
        "행복한큰별이", #P18, 44
        "뚜뚜치맘", #P19, 43
        "네찌늦둥이", #P20, 65, P19이랑 같은 사람 -> 서로 다른 임신기간, 아이디 바꿈
        "동욱경모맘", #P21, 103
        "큰기린" #P22, 22
        ]
data_path = "./data/"
xl_filename = "./articles.xlsx"

if __name__ == "__main__" :
    print('start')
    # input json files
    file_list = os.listdir(data_path)
    # output excel file
    wb = openpyxl.load_workbook(filename=xl_filename)

    for i in range(23) :
        pid = 'P'+str(i)
        nickname = nicknames[i]
        # create sheet
        ws = wb.create_sheet(nickname)
        print(pid, ': ', nickname)
        target_list = [f for f in file_list if pid in f]

        index = 2
        for aid, target in enumerate(target_list) :
            with open(data_path+target) as json_file:
                json_data = json.load(json_file)

                sentences = json_data['contents']

                #print(json_data)
                
                # sid: sid
                for sid, sentence in enumerate(sentences) :
                    # pid: pid
                    v_sentence = sentence.replace("\u200B", "").strip()
                    if v_sentence != '' and v_sentence != '\n' :
                        #print(v_sentence)
                        ws['A'+str(index)] = pid
                        # aid: aid
                        ws['B'+str(index)] = aid
                        ws['C'+str(index)] = sid
                        ws['D'+str(index)] = v_sentence
                        index += 1

    wb.save(filename=xl_filename)


        
        
